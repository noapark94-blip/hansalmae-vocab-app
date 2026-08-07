#!/usr/bin/env python3
"""기존 한살매 Excel/Google Sheets 내보내기를 Supabase로 이관한다.

사용법:
  pip install -r tools/requirements.txt
  python tools/import_excel.py 기존파일.xlsx --dry-run
  python tools/import_excel.py 기존파일.xlsx
"""
import argparse, json, os, re, sys
from pathlib import Path
from openpyxl import load_workbook

try:
    import requests
except ImportError:
    requests = None

def text(v): return "" if v is None else str(v).strip()
def number(v, default=0):
    try: return int(float(v))
    except (TypeError, ValueError): return default
def truth(v): return text(v).lower() in {"true","1","y","yes","사용","사용함","활성","o"}
def chunks(items, n=500):
    for i in range(0, len(items), n): yield items[i:i+n]

class Supabase:
    def __init__(self, base, key, dry=False):
        self.base=base.rstrip('/'); self.key=key; self.dry=dry
        self.h={"apikey":key,"Authorization":f"Bearer {key}","Content-Type":"application/json","Prefer":"return=representation,resolution=merge-duplicates"}
    def get(self, table, params):
        r=requests.get(f"{self.base}/rest/v1/{table}",headers=self.h,params=params,timeout=60); r.raise_for_status(); return r.json()
    def post(self, table, rows, conflict=None):
        if not rows:return []
        if self.dry:return rows
        params={"on_conflict":conflict} if conflict else {}
        r=requests.post(f"{self.base}/rest/v1/{table}",headers=self.h,params=params,json=rows,timeout=120)
        if not r.ok: raise RuntimeError(f"{table}: {r.status_code} {r.text}")
        return r.json() if r.text else []
    def create_user(self, student_id, password, name, grade, year, enabled):
        if self.dry:return {"id":f"dry-{student_id}"}
        payload={"email":f"{re.sub(r'[^a-z0-9._-]','-',student_id.lower())}@student.hansalmae.local","password":password,"email_confirm":True,"user_metadata":{"display_name":name}}
        r=requests.post(f"{self.base}/auth/v1/admin/users",headers=self.h,json=payload,timeout=60)
        if r.status_code==422:
            existing=self.get("profiles",{"select":"id","student_id":f"eq.{student_id.lower()}"})
            if existing:return {"id":existing[0]["id"]}
        r.raise_for_status(); user=r.json().get("user",r.json())
        self.post("profiles",[{"id":user["id"],"student_id":student_id.lower(),"display_name":name,"base_grade":grade,"base_year":year,"enabled":enabled,"role":"student"}],"student_id")
        self.post("student_experience",[{"user_id":user["id"]}],"user_id")
        existing=self.get("vocabulary_books",{"select":"id","user_id":f"eq.{user['id']}","is_default":"eq.true"})
        if not existing:self.post("vocabulary_books",[{"user_id":user["id"],"name":"기본 단어장","is_default":True}])
        return user

def rows(ws):
    values=list(ws.iter_rows(values_only=True))
    if not values:return []
    headers=[text(x) for x in values[0]]
    return [{headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i]} for row in values[1:] if any(text(x) for x in row)]

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("xlsx",type=Path); ap.add_argument("--dry-run",action="store_true"); a=ap.parse_args()
    base=os.getenv("SUPABASE_URL",""); key=os.getenv("SUPABASE_SERVICE_ROLE_KEY","")
    if not a.dry_run and (not base or not key):sys.exit("SUPABASE_URL과 SUPABASE_SERVICE_ROLE_KEY 환경변수가 필요합니다.")
    if not a.dry_run and requests is None:sys.exit("requests가 없습니다. 먼저 pip install -r tools/requirements.txt 를 실행하세요.")
    sb=Supabase(base or "https://dry-run.invalid",key or "dry",a.dry_run); wb=load_workbook(a.xlsx,data_only=True,read_only=True)
    report={"word_sets":0,"words":0,"students":0,"warnings":[]}
    for ws in wb.worksheets:
        if "단어DB" not in ws.title:continue
        set_rows=sb.post("word_sets",[{"name":ws.title,"slug":f"word-set-{report['word_sets']+1}","sort_order":report["word_sets"]}],"name")
        set_id=(set_rows[0].get("id") if set_rows else None)
        if not set_id and not a.dry_run:
            found=sb.get("word_sets",{"select":"id","name":f"eq.{ws.title}"}); set_id=found[0]["id"]
        if a.dry_run:set_id=f"dry-set-{report['word_sets']}"
        data=[]
        for i,r in enumerate(rows(ws)):
            word=text(r.get("영어")); meaning=text(r.get("뜻"))
            if not word or not meaning:continue
            day_text=text(r.get("Day")); match=re.search(r"\d+",day_text)
            if not match:continue
            data.append({"word_set_id":set_id,"day":int(match.group()),"word":word,"meaning":meaning,"example":text(r.get("예문")),"translation":text(r.get("예문 해석")),"example_answer":text(r.get("예문 정답")),"sort_order":i})
        # 같은 DB/Day/영어 조합이 Excel 안에 중복된 경우 Postgres upsert가
        # 한 명령 안에서 같은 행을 두 번 갱신하지 않도록 마지막 항목만 남긴다.
        deduplicated = {}
        for item in data:
            deduplicated[(item["word_set_id"], item["day"], item["word"])] = item
        data = list(deduplicated.values())
        for part in chunks(data):sb.post("words",part,"word_set_id,day,word")
        report["word_sets"]+=1;report["words"]+=len(data)
    if "학생계정" in wb.sheetnames:
        for r in rows(wb["학생계정"]):
            sid=text(r.get("학생ID")); password=text(r.get("비밀번호")); name=text(r.get("학생이름"))
            if not sid or not password or not name:continue
            try:sb.create_user(sid,password,name,text(r.get("기준학년")),number(r.get("기준연도")),truth(r.get("사용여부")));report["students"]+=1
            except Exception as e:report["warnings"].append(f"학생 {sid}: {e}")
    out=Path(__file__).with_name("import-report.json");out.write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps(report,ensure_ascii=False,indent=2));print(f"보고서: {out}")
if __name__=="__main__":main()
